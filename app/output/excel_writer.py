from pathlib import Path
import pandas as pd
from openpyxl.utils import get_column_letter

from app.utils.logger import setup_logger
from app.config.settings import Settings

logger = setup_logger(Settings.LOG_LEVEL)


def ensure_output_dir():
    output_dir = Path(Settings.OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def write_excel(df: pd.DataFrame):
    output_dir = ensure_output_dir()
    excel_path = output_dir / "invoices.xlsx"

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Invoices", index=False)

        sheet = writer.book["Invoices"]

        # Auto-size columns safely (no MergedCell issues)
        for i, col in enumerate(sheet.iter_cols(values_only=True), start=1):
            max_length = max(
                (len(str(cell)) for cell in col if cell is not None),
                default=10
            )

            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = max_length + 2

    logger.info(f"Excel written to {excel_path}")
    return excel_path
